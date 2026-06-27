import time
from datetime import datetime, timezone
from threading import Lock, Thread, Event
import hashlib
import re

import log

from lib.config import XLXS_PATH, STOCK_CODES, IDX_FLUSH_EVERY, SOURCE_INFO, CONFIG_KEYWORDS_DATA
from lib.models import (
    SOURCE_INDEX_HEADERS,
    NEWS_RAW_HEADERS,
    CONFIG_SOURCES_HEADERS,
    CONFIG_KEYWORDS_HEADERS,
    CRAWL_LOG_HEADERS,
    DAILY_SUMMARY_HEADERS,
    DATA_QUALITY_CHECK_HEADERS,
    README_HEADERS,
    README_ROWS,
)

SOURCE_INDEX_NAME = "SOURCE_INDEX"
NEWS_RAW_NAME = "NEWS_RAW"
CONFIG_SOURCES_NAME = "CONFIG_SOURCES"
CONFIG_KEYWORDS_NAME = "CONFIG_KEYWORDS"
CRAWL_LOG_NAME = "CRAWL_LOG"
DAILY_SUMMARY_NAME = "DAILY_SUMMARY"
DATA_QUALITY_NAME = "DATA_QUALITY_CHECK"
README_NAME = "README"
COMPANY_INFO_NAME = "COMPANY_INFO"
NGUON_TIN_NAME = "SOURCE_LIST"


_idx_accumulator: list[list] = []
_idx_links: set[str] = set()
_idx_flush_counter: int = 0
_idx_lock = Lock()

_writer_thread: Thread | None = None
_writer_stop = Event()
_writer_flush_request = Event()
_writer_lock = Lock()
_writer_pending = 0

_raw_accumulator: list[list] = []
_raw_hashes: set[str] = set()
_raw_flush_counter: int = 0
_raw_lock = Lock()
RAW_FLUSH_EVERY = 200


def _xl_styles():
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    return {
        "hdr_fill": PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid"),
        "hdr_font": Font(name="Calibri", bold=True, color="FFFFFF", size=11),
        "body_font": Font(name="Calibri", size=11),
        "thin": Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")),
        "wrap": Alignment(wrap_text=True),
        "center_wrap": Alignment(horizontal="center", vertical="center", wrap_text=True),
    }


def _write_sheet(ws, headers, rows_data, col_widths, styles, freeze="A2"):
    s = styles
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = s["hdr_fill"]; c.font = s["hdr_font"]; c.border = s["thin"]; c.alignment = s["center_wrap"]
    for ri, row in enumerate(rows_data, 2):
        for ci, v in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font = s["body_font"]; c.border = s["thin"]; c.alignment = s["wrap"]
    if col_widths:
        from openpyxl.utils import get_column_letter
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    last_col = __import__("openpyxl").utils.get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}{len(rows_data) + 1}"
    ws.freeze_panes = freeze


def _ensure_sheet_exists(wb, name: str, headers: list[str], s: dict) -> None:
    if name not in wb.sheetnames:
        ws = wb.create_sheet(name)
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = s["hdr_font"]; c.fill = s["hdr_fill"]; c.border = s["thin"]; c.alignment = s["center_wrap"]


def _make_sheet_with_rows(wb, name, headers, rows, col_widths, s, position=None):
    if name not in wb.sheetnames:
        ws = wb.create_sheet(name, position) if position is not None else wb.create_sheet(name)
        _write_sheet(ws, headers, rows, col_widths, s)
    else:
        _ensure_sheet_exists(wb, name, headers, s)


def init_workbook(output_path: str = ""):
    if not output_path:
        output_path = XLXS_PATH
    import openpyxl as xl
    import os

    exists = os.path.exists(output_path)
    s = _xl_styles()

    if exists:
        wb = xl.load_workbook(output_path)
        _ensure_sheet_exists(wb, SOURCE_INDEX_NAME, SOURCE_INDEX_HEADERS, s)
        _init_accumulator(wb)
        _ensure_new_sheets(wb, s)
        return wb

    wb = xl.Workbook()

    ws1 = wb.active
    ws1.title = COMPANY_INFO_NAME
    from lib.config import COMPANY_INFO
    fields = ["Mã CK", "Tên đầy đủ", "Tên tiếng Anh", "Tên cũ", "Sàn", "Ngành", "Website", "Thành lập", "Nhân sự", "Kinh doanh chính", "Từ khóa"]
    rows = []
    for ci in COMPANY_INFO:
        rows.append([ci["code"], ci["full_name"], ci["english_name"], ci["former_name"], ci["exchange"], ci["industry"], ci["website"], ci["established"], ci["employees"], ci["business"], ci["keywords"]])
    _write_sheet(ws1, fields, rows, [8, 40, 40, 35, 8, 25, 30, 10, 10, 60, 40], s)

    _make_sheet_with_rows(wb, NGUON_TIN_NAME,
        ["STT", "Tên trang", "Domain", "Chuyên mục", "Loại", "RSS", "Số bài", "Ghi chú"],
        [[i + 1, si["name"], si["domain"], si["categories"], si["type"], si["rss"], si["articles"], si["note"]] for i, si in enumerate(SOURCE_INFO)],
        [5, 18, 22, 20, 22, 10, 14, 50], s, position=1)

    ws_idx = wb.create_sheet(SOURCE_INDEX_NAME)
    for col, h in enumerate(SOURCE_INDEX_HEADERS, 1):
        c = ws_idx.cell(row=1, column=col, value=h)
        c.font = s["hdr_font"]; c.fill = s["hdr_fill"]; c.border = s["thin"]; c.alignment = s["center_wrap"]

    _create_new_sheets(wb, s)

    wb.save(output_path)
    return wb


def _create_new_sheets(wb, s):
    _make_sheet_with_rows(wb, NEWS_RAW_NAME, NEWS_RAW_HEADERS, [],
        [10, 50, 30, 60, 14, 20, 50, 20, 20, 30, 15, 20, 10, 12, 40, 15, 20, 30], s)
    src_rows = []
    for si in SOURCE_INFO:
        methods = []
        if si.get("has_rss"):
            methods.append("rss")
        if si.get("has_api"):
            methods.append("api")
        if not methods:
            methods.append("manual")
        src_rows.append([
            si["name"].upper().replace(" ", "_").replace("&", "").replace(".", ""),
            si["name"],
            "https://" + si["domain"],
            "",
            "",
            si["categories"],
            si["type"],
            "yes" if si.get("has_rss") else "no",
            "yes" if si.get("has_api") else "no",
            "yes",
            "/".join(methods),
            si.get("articles", ""),
            si.get("note", ""),
        ])
    _make_sheet_with_rows(wb, CONFIG_SOURCES_NAME, CONFIG_SOURCES_HEADERS, src_rows,
        [12, 20, 30, 35, 35, 25, 15, 8, 8, 10, 15, 10, 40], s)
    kw_rows = []
    for kw in CONFIG_KEYWORDS_DATA:
        kw_rows.append([
            kw["keyword_id"],
            kw["keyword"],
            kw["industry_group"],
            kw["related_tickers"],
            kw["event_type_suggestion"],
            kw["priority"],
            kw["note"],
        ])
    _make_sheet_with_rows(wb, CONFIG_KEYWORDS_NAME, CONFIG_KEYWORDS_HEADERS, kw_rows,
        [10, 40, 20, 20, 20, 10, 40], s)
    _make_sheet_with_rows(wb, CRAWL_LOG_NAME, CRAWL_LOG_HEADERS, [],
        [8, 14, 12, 20, 20, 14, 14, 12, 12, 10, 8, 10, 40], s)
    _make_sheet_with_rows(wb, DAILY_SUMMARY_NAME, DAILY_SUMMARY_HEADERS, [],
        [12, 14, 16, 20, 18, 18, 14, 14, 30, 30], s)
    _make_sheet_with_rows(wb, DATA_QUALITY_NAME, DATA_QUALITY_CHECK_HEADERS, [],
        [8, 14, 15, 12, 12, 10, 18, 12, 12, 18, 14, 40], s)
    _make_sheet_with_rows(wb, README_NAME, README_HEADERS, README_ROWS,
        [20, 50], s)


def _ensure_new_sheets(wb, s):
    existing = set(wb.sheetnames)
    if README_NAME not in existing:
        _create_new_sheets(wb, s)
    else:
        _ensure_sheet_exists(wb, NEWS_RAW_NAME, NEWS_RAW_HEADERS, s)
        _ensure_sheet_exists(wb, CONFIG_SOURCES_NAME, CONFIG_SOURCES_HEADERS, s)
        _ensure_sheet_exists(wb, CONFIG_KEYWORDS_NAME, CONFIG_KEYWORDS_HEADERS, s)
        _ensure_sheet_exists(wb, CRAWL_LOG_NAME, CRAWL_LOG_HEADERS, s)
        _ensure_sheet_exists(wb, DAILY_SUMMARY_NAME, DAILY_SUMMARY_HEADERS, s)
        _ensure_sheet_exists(wb, DATA_QUALITY_NAME, DATA_QUALITY_CHECK_HEADERS, s)
        _ensure_sheet_exists(wb, README_NAME, README_HEADERS, s)
        if README_NAME in wb.sheetnames and wb[README_NAME].max_row < 2:
            ws = wb[README_NAME]
            for ri, row in enumerate(README_ROWS, 2):
                for ci, v in enumerate(row, 1):
                    ws.cell(ri, ci, v)


def _init_accumulator(wb):
    global _idx_accumulator, _idx_links, _idx_flush_counter
    _idx_accumulator = []
    _idx_links = set()
    _idx_flush_counter = 0
    if SOURCE_INDEX_NAME in wb.sheetnames:
        ws = wb[SOURCE_INDEX_NAME]
        for r in range(2, ws.max_row + 1):
            v = ws.cell(r, 1).value
            if v:
                _idx_links.add(str(v).strip())


def _init_raw_accumulator(wb):
    global _raw_accumulator, _raw_hashes, _raw_flush_counter
    _raw_accumulator = []
    _raw_hashes = set()
    _raw_flush_counter = 0
    if NEWS_RAW_NAME in wb.sheetnames:
        ws = wb[NEWS_RAW_NAME]
        for r in range(2, ws.max_row + 1):
            v = ws.cell(r, 1).value
            if v:
                _raw_hashes.add(str(v).strip())


def _make_news_id(source: str, pub_date: str, url: str) -> str:
    prefix = source.split(" ")[0].upper()[:8]
    date_part = pub_date[:10].replace("-", "") if pub_date else "00000000"
    h = hashlib.md5(url.encode()).hexdigest()[:6].upper()
    return f"{prefix}_{date_part}_{h}"


def accumulate_raw_entry(entry: dict) -> int:
    global _raw_accumulator, _raw_hashes, _raw_flush_counter
    with _raw_lock:
        url = entry.get("url", "")
        if not url:
            return 0
        h = hashlib.md5(url.encode()).hexdigest()[:6].upper()
        if h in _raw_hashes:
            return 0
        _raw_hashes.add(h)
        news_id = _make_news_id(entry.get("source", ""), entry.get("published_date", ""), url)
        row = [
            news_id,
            entry.get("title", ""),
            entry.get("summary", ""),
            entry.get("content", ""),
            entry.get("published_date", ""),
            entry.get("source", ""),
            url,
            entry.get("industry_group", ""),
            entry.get("tickers", ""),
            entry.get("keywords", ""),
            entry.get("event_type", ""),
            datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S"),
            h,
            entry.get("crawl_status", "new"),
            entry.get("error_message", ""),
            "",
            "",
            "",
        ]
        _raw_accumulator.append(row)
        _raw_flush_counter += 1
    return 1


def flush_raw_accumulator(wb, output_path: str, s: dict) -> int:
    global _raw_accumulator, _raw_flush_counter
    with _raw_lock:
        if not _raw_accumulator:
            return 0
        _ensure_sheet_exists(wb, NEWS_RAW_NAME, NEWS_RAW_HEADERS, s)
        ws = wb[NEWS_RAW_NAME]
        for row in _raw_accumulator:
            ws.append(row)
        count = len(_raw_accumulator)
        _raw_accumulator = []
        _raw_flush_counter = 0
        wb.save(output_path)
    return count


def maybe_flush_raw(wb, output_path: str, s: dict) -> int:
    with _raw_lock:
        if _raw_flush_counter >= RAW_FLUSH_EVERY:
            should = True
        else:
            should = False
    if should:
        return flush_raw_accumulator(wb, output_path, s)
    return 0


def _accumulate_entry(entry: dict) -> int:
    global _idx_accumulator, _idx_links, _idx_flush_counter
    with _idx_lock:
        link = entry.get("link", "")
        if not link or link in _idx_links:
            return 0
        _idx_links.add(link)
        row = [
            link,
            entry.get("source", ""),
            entry.get("category", ""),
            entry.get("title", ""),
            entry.get("datetime_public", ""),
            entry.get("status", "new"),
        ]
        _idx_accumulator.append(row)
        _idx_flush_counter += 1
    return 1


def _flush_accumulator(wb, output_path: str, s: dict) -> int:
    global _idx_accumulator, _idx_flush_counter
    with _idx_lock:
        if not _idx_accumulator:
            return 0
        _ensure_sheet_exists(wb, SOURCE_INDEX_NAME, SOURCE_INDEX_HEADERS, s)
        ws = wb[SOURCE_INDEX_NAME]
        for row in _idx_accumulator:
            ws.append(row)
        count = len(_idx_accumulator)
        _idx_accumulator = []
        _idx_flush_counter = 0
        wb.save(output_path)
    return count


def _maybe_flush(wb, output_path: str, s: dict) -> int:
    with _idx_lock:
        if _idx_flush_counter >= IDX_FLUSH_EVERY:
            _writer_flush_request.set()
            return 1
    return 0


def start_background_writer(wb, output_path: str, s):
    global _writer_thread
    if _writer_thread is not None:
        return
    _writer_stop.clear()
    _writer_flush_request.clear()
    _writer_thread = Thread(target=_writer_loop, args=(wb, output_path, s), daemon=True)
    _writer_thread.start()


def _writer_loop(wb, output_path, s):
    log.info("Writer thread started")
    last_flush_size = 0
    while not _writer_stop.is_set():
        if _writer_flush_request.is_set():
            _writer_flush_request.clear()
            n = _flush_accumulator(wb, output_path, s)
            if n:
                log.info(f"Writer flushed {n} entries")
        _writer_flush_request.wait(0.5)
    log.info("Writer thread stopped")


def stop_background_writer(wb, output_path, s):
    global _writer_thread
    if _writer_thread is None:
        return
    _writer_flush_request.set()
    _writer_stop.set()
    _writer_thread.join(timeout=30)
    _flush_accumulator(wb, output_path, s)
    _writer_thread = None

def get_source_index_rows(wb) -> list[dict]:
    if SOURCE_INDEX_NAME not in wb.sheetnames:
        return []
    ws = wb[SOURCE_INDEX_NAME]
    rows_data = list(ws.iter_rows(values_only=True))
    if len(rows_data) < 2:
        return []
    headers = rows_data[0]
    rows = []
    for row_values in rows_data[1:]:
        row_dict = {}
        for idx, val in enumerate(row_values):
            if idx < len(headers) and headers[idx] is not None:
                row_dict[headers[idx]] = val
        rows.append(row_dict)
    return rows



def append_crawl_log(wb, output_path: str, s: dict, entry: dict) -> None:
    _ensure_sheet_exists(wb, CRAWL_LOG_NAME, CRAWL_LOG_HEADERS, s)
    ws = wb[CRAWL_LOG_NAME]
    next_id = ws.max_row
    row = [
        f"CL{next_id:05d}",
        entry.get("log_date", datetime.now(timezone.utc).strftime("%Y-%m-%d")),
        entry.get("member", ""),
        entry.get("source", ""),
        entry.get("keyword_group", ""),
        entry.get("date_range_from", ""),
        entry.get("date_range_to", ""),
        entry.get("records_found", 0),
        entry.get("records_added", 0),
        entry.get("duplicates", 0),
        entry.get("failed", 0),
        entry.get("status", "done"),
        entry.get("note", ""),
    ]
    ws.append(row)
    wb.save(output_path)


def update_daily_summary(wb, output_path: str, s: dict, entry: dict) -> None:
    _ensure_sheet_exists(wb, DAILY_SUMMARY_NAME, DAILY_SUMMARY_HEADERS, s)
    ws = wb[DAILY_SUMMARY_NAME]
    row = [
        entry.get("date", datetime.now(timezone.utc).strftime("%Y-%m-%d")),
        entry.get("total_records", 0),
        entry.get("new_records_today", 0),
        entry.get("sources_updated", ""),
        entry.get("date_coverage_from", ""),
        entry.get("date_coverage_to", ""),
        entry.get("missing_days", ""),
        entry.get("duplicate_count", 0),
        entry.get("issue_status", ""),
        entry.get("next_action", ""),
    ]
    ws.append(row)
    wb.save(output_path)
